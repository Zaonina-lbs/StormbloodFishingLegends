"""
数据加载模块 - 从 Excel 文件读取鱼基础数据和鱼饵数据并导入数据库
"""

import openpyxl
import os
try:
    from .database import get_db
except ImportError:
    from database import get_db


def load_fish_data_from_excel(filepath="fish_data.xlsx"):
    """从 fish_data.xlsx 加载鱼基础数据"""
    if not os.path.isabs(filepath):
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filepath)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    fish_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        fish = {
            "name": str(row[0]).strip() if row[0] else "",
            "region": str(row[1]).strip() if row[1] else "",
            "fishing_ground": str(row[2]).strip() if row[2] else "",
            "bait": str(row[3]).strip() if row[3] else "",
            "weather": str(row[4]).strip() if row[4] else "",
            "fish_type": str(row[5]).strip() if row[5] else "普通鱼",
            "min_size": float(row[6]) if row[6] else 0.0,
            "min_big_size": float(row[7]) if row[7] else 0.0,
            "max_size": float(row[8]) if row[8] else 0.0,
            "base_value": float(row[9]) if row[9] else 50.0,
        }
        fish_list.append(fish)
    wb.close()
    return fish_list


def load_lure_data_from_excel(filepath="fishingLure.xlsx"):
    """从 fishingLure.xlsx 加载鱼饵数据"""
    if not os.path.isabs(filepath):
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filepath)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    lure_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        lure = {
            "name": str(row[0]).strip() if row[0] else "",
            "sellable": str(row[1]).strip() == "可以" if row[1] else False,
            "price": int(row[2]) if row[2] else None,
        }
        lure_list.append(lure)
    wb.close()
    return lure_list


def export_fish_to_yaml(filepath="fish_data.yaml"):
    """将 Excel 鱼数据导出为 YAML 配置文件"""
    fish_list = load_fish_data_from_excel()
    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), filepath)
    with open(output, "w", encoding="utf-8") as f:
        f.write("# 鱼基础数据配置文件\n")
        f.write("# 可直接修改此文件，然后通过 hot_reload 热更新\n")
        f.write("# bait 为空表示任何鱼饵都能上钩\n\n")
        f.write("fishes:\n")
        for fish in fish_list:
            f.write(f'  - name: "{fish["name"]}"\n')
            f.write(f'    region: "{fish["region"]}"\n')
            f.write(f'    fishing_ground: "{fish["fishing_ground"]}"\n')
            f.write(f'    bait: "{fish["bait"]}"\n')
            f.write(f'    weather: "{fish["weather"]}"\n')
            f.write(f'    fish_type: "{fish["fish_type"]}"\n')
            f.write(f"    min_size: {fish['min_size']}\n")
            f.write(f"    min_big_size: {fish['min_big_size']}\n")
            f.write(f"    max_size: {fish['max_size']}\n")
            f.write(f"    base_value: {int(fish['base_value'])}\n")
    print(f"✅ 鱼数据已导出到 {output}")


def export_lure_to_yaml(filepath="lure_data.yaml"):
    """将 Excel 鱼饵数据导出为 YAML 配置文件"""
    lure_list = load_lure_data_from_excel()
    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), filepath)
    with open(output, "w", encoding="utf-8") as f:
        f.write("# 鱼饵数据配置文件\n")
        f.write("# 可直接修改此文件，然后通过 hot_reload 热更新\n\n")
        f.write("lures:\n")
        for lure in lure_list:
            price_str = lure["price"] if lure["price"] is not None else "null"
            f.write(f'  - name: "{lure["name"]}"\n')
            f.write(f"    sellable: {'true' if lure['sellable'] else 'false'}\n")
            f.write(f"    price: {price_str}\n")
    print(f"✅ 鱼饵数据已导出到 {output}")


def init_game_data():
    """初始化游戏数据：从 Excel 加载并导入数据库"""
    db = get_db()
    db.init_database()
    print("正在加载鱼基础数据...")
    fish_list = load_fish_data_from_excel()
    db.import_fish_data(fish_list)
    print(f"已导入 {len(fish_list)} 条鱼基础数据")
    print("正在加载鱼饵数据...")
    lure_list = load_lure_data_from_excel()
    db.import_lure_data(lure_list)
    print(f"已导入 {len(lure_list)} 条鱼饵数据")
    print("正在生成天气数据...")
    from datetime import date, timedelta

    today = date.today()
    for i in range(4):
        d = (today + timedelta(days=i)).isoformat()
        for slot in (0, 1):
            db.generate_weather(d, slot)
    print("天气数据已生成")
    # 自动导出 YAML 配置文件
    export_fish_to_yaml()
    export_lure_to_yaml()
    print("游戏数据初始化完成！")


if __name__ == "__main__":
    init_game_data()
